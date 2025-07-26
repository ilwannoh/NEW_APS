import pandas as pd
import openpyxl
from pathlib import Path
from typing import Dict, List, Optional, Any, Union
import logging
from datetime import datetime


class ExcelHandler:
    """
    엑셀 파일 처리를 위한 유틸리티 클래스
    """
    
    def __init__(self, logger: Optional[logging.Logger] = None):
        self.logger = logger or logging.getLogger(__name__)
    
    def read_excel_file(self, file_path: Union[str, Path], sheet_name: Optional[str] = None) -> Dict[str, pd.DataFrame]:
        """
        엑셀 파일을 읽어서 DataFrame 딕셔너리로 반환
        
        Args:
            file_path: 엑셀 파일 경로
            sheet_name: 특정 시트명 (None이면 모든 시트)
        
        Returns:
            {시트명: DataFrame} 형태의 딕셔너리
        """
        try:
            file_path = Path(file_path)
            if not file_path.exists():
                raise FileNotFoundError(f"파일을 찾을 수 없습니다: {file_path}")
            
            self.logger.info(f"엑셀 파일 읽기 시작: {file_path}")
            
            if sheet_name:
                # 특정 시트만 읽기
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                result = {sheet_name: df}
            else:
                # 모든 시트 읽기
                excel_file = pd.ExcelFile(file_path)
                result = {}
                for sheet in excel_file.sheet_names:
                    try:
                        df = pd.read_excel(file_path, sheet_name=sheet)
                        result[sheet] = df
                        self.logger.debug(f"시트 '{sheet}' 읽기 완료: {len(df)}행, {len(df.columns)}열")
                    except Exception as e:
                        self.logger.warning(f"시트 '{sheet}' 읽기 실패: {e}")
                        continue
            
            self.logger.info(f"엑셀 파일 읽기 완료: {len(result)}개 시트")
            return result
            
        except Exception as e:
            self.logger.error(f"엑셀 파일 읽기 오류: {e}")
            raise
    
    def write_excel_file(self, data: Dict[str, pd.DataFrame], file_path: Union[str, Path]) -> bool:
        """
        DataFrame 딕셔너리를 엑셀 파일로 저장
        
        Args:
            data: {시트명: DataFrame} 형태의 딕셔너리
            file_path: 저장할 파일 경로
        
        Returns:
            성공 여부
        """
        try:
            file_path = Path(file_path)
            
            # 디렉토리 생성
            file_path.parent.mkdir(parents=True, exist_ok=True)
            
            self.logger.info(f"엑셀 파일 쓰기 시작: {file_path}")
            
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                for sheet_name, df in data.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    self.logger.debug(f"시트 '{sheet_name}' 쓰기 완료: {len(df)}행")
            
            self.logger.info(f"엑셀 파일 쓰기 완료: {len(data)}개 시트")
            return True
            
        except Exception as e:
            self.logger.error(f"엑셀 파일 쓰기 오류: {e}")
            return False
    
    def get_sheet_info(self, file_path: Union[str, Path]) -> Dict[str, Dict[str, Any]]:
        """
        엑셀 파일의 시트 정보를 반환
        
        Args:
            file_path: 엑셀 파일 경로
        
        Returns:
            시트별 정보 딕셔너리
        """
        try:
            file_path = Path(file_path)
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            sheet_info = {}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # 시트의 사용된 범위 확인
                max_row = ws.max_row
                max_col = ws.max_column
                
                # 실제 데이터가 있는 행 수 계산
                actual_rows = 0
                for row in range(1, max_row + 1):
                    if any(ws.cell(row, col).value is not None for col in range(1, max_col + 1)):
                        actual_rows = row
                
                sheet_info[sheet_name] = {
                    'max_row': max_row,
                    'max_column': max_col,
                    'actual_rows': actual_rows,
                    'has_data': actual_rows > 0
                }
            
            wb.close()
            return sheet_info
            
        except Exception as e:
            self.logger.error(f"시트 정보 조회 오류: {e}")
            return {}
    
    def parse_plan_sheet(self, df: pd.DataFrame) -> List[Dict[str, Any]]:
        """
        plan.xlsx의 주간생산계획 시트를 파싱
        
        Args:
            df: 주간생산계획 DataFrame
        
        Returns:
            파싱된 생산계획 데이터 리스트
        """
        try:
            parsed_data = []
            
            # 컬럼명 정리 (공백 제거)
            df.columns = df.columns.str.strip()
            
            for idx, row in df.iterrows():
                if pd.isna(row.iloc[0]):  # 첫 번째 컬럼이 비어있으면 스킵
                    continue
                
                plan_data = {
                    'row_index': idx,
                    'product_code': str(row.iloc[0]).strip() if not pd.isna(row.iloc[0]) else '',
                    'planned_quantity': self._safe_numeric_convert(row.iloc[1]) if len(row) > 1 else 0,
                    'priority': self._safe_numeric_convert(row.iloc[2]) if len(row) > 2 else 5,
                    'due_date': self._safe_date_convert(row.iloc[3]) if len(row) > 3 else None,
                    'raw_data': row.to_dict()
                }
                
                parsed_data.append(plan_data)
            
            self.logger.info(f"생산계획 파싱 완료: {len(parsed_data)}개 항목")
            return parsed_data
            
        except Exception as e:
            self.logger.error(f"생산계획 파싱 오류: {e}")
            return []
    
    def parse_basic_sheet(self, df: pd.DataFrame) -> List[Dict[str, Any]]:
        """
        basic.xlsx의 제품기준정보 시트를 파싱
        
        Args:
            df: 제품기준정보 DataFrame
        
        Returns:
            파싱된 제품정보 데이터 리스트
        """
        try:
            parsed_data = []
            
            # 컬럼명 정리
            df.columns = df.columns.str.strip()
            
            for idx, row in df.iterrows():
                if pd.isna(row.iloc[0]):
                    continue
                
                product_data = {
                    'row_index': idx,
                    'product_code': str(row.iloc[0]).strip() if not pd.isna(row.iloc[0]) else '',
                    'product_name': str(row.iloc[1]).strip() if len(row) > 1 and not pd.isna(row.iloc[1]) else '',
                    'specification': str(row.iloc[2]).strip() if len(row) > 2 and not pd.isna(row.iloc[2]) else '',
                    'unit': str(row.iloc[3]).strip() if len(row) > 3 and not pd.isna(row.iloc[3]) else '',
                    'standard_time': self._safe_numeric_convert(row.iloc[4]) if len(row) > 4 else 0,
                    'raw_data': row.to_dict()
                }
                
                parsed_data.append(product_data)
            
            self.logger.info(f"제품정보 파싱 완료: {len(parsed_data)}개 항목")
            return parsed_data
            
        except Exception as e:
            self.logger.error(f"제품정보 파싱 오류: {e}")
            return []
    
    def _safe_numeric_convert(self, value: Any) -> float:
        """
        안전한 숫자 변환
        """
        try:
            if pd.isna(value):
                return 0.0
            if isinstance(value, (int, float)):
                return float(value)
            if isinstance(value, str):
                # 쉼표 제거 후 변환
                cleaned = value.replace(',', '').strip()
                return float(cleaned) if cleaned else 0.0
            return 0.0
        except:
            return 0.0
    
    def _safe_date_convert(self, value: Any) -> Optional[datetime]:
        """
        안전한 날짜 변환
        """
        try:
            if pd.isna(value):
                return None
            if isinstance(value, datetime):
                return value
            if isinstance(value, str):
                # 다양한 날짜 형식 시도
                for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%m/%d/%Y', '%d/%m/%Y']:
                    try:
                        return datetime.strptime(value.strip(), fmt)
                    except:
                        continue
            return None
        except:
            return None
    
    def validate_excel_structure(self, file_path: Union[str, Path], required_sheets: List[str]) -> Dict[str, bool]:
        """
        엑셀 파일 구조 검증
        
        Args:
            file_path: 엑셀 파일 경로
            required_sheets: 필요한 시트 목록
        
        Returns:
            시트별 존재 여부
        """
        try:
            file_path = Path(file_path)
            if not file_path.exists():
                return {sheet: False for sheet in required_sheets}
            
            wb = openpyxl.load_workbook(file_path)
            existing_sheets = wb.sheetnames
            wb.close()
            
            result = {}
            for sheet in required_sheets:
                result[sheet] = sheet in existing_sheets
            
            return result
            
        except Exception as e:
            self.logger.error(f"엑셀 구조 검증 오류: {e}")
            return {sheet: False for sheet in required_sheets}


# 편의 함수들
def read_excel(file_path: Union[str, Path], sheet_name: Optional[str] = None) -> Dict[str, pd.DataFrame]:
    """
    간단한 엑셀 읽기 함수
    """
    handler = ExcelHandler()
    return handler.read_excel_file(file_path, sheet_name)


def write_excel(data: Dict[str, pd.DataFrame], file_path: Union[str, Path]) -> bool:
    """
    간단한 엑셀 쓰기 함수
    """
    handler = ExcelHandler()
    return handler.write_excel_file(data, file_path)


def get_excel_info(file_path: Union[str, Path]) -> Dict[str, Dict[str, Any]]:
    """
    간단한 엑셀 정보 조회 함수
    """
    handler = ExcelHandler()
    return handler.get_sheet_info(file_path)

