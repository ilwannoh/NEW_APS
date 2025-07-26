from typing import Any, List, Dict, Optional, Union, Callable
from datetime import datetime, date
import re
import logging
from pathlib import Path
import pandas as pd


class ValidationError(Exception):
    """데이터 검증 실패 예외"""
    pass


class ValidationResult:
    """
    검증 결과를 나타내는 클래스
    """
    
    def __init__(self, is_valid: bool = True, errors: List[str] = None, warnings: List[str] = None):
        self.is_valid = is_valid
        self.errors = errors or []
        self.warnings = warnings or []
    
    def add_error(self, message: str) -> None:
        """오류 메시지 추가"""
        self.errors.append(message)
        self.is_valid = False
    
    def add_warning(self, message: str) -> None:
        """경고 메시지 추가"""
        self.warnings.append(message)
    
    def merge(self, other: 'ValidationResult') -> None:
        """다른 검증 결과와 병합"""
        self.errors.extend(other.errors)
        self.warnings.extend(other.warnings)
        if not other.is_valid:
            self.is_valid = False
    
    def get_summary(self) -> Dict[str, Any]:
        """검증 결과 요약 반환"""
        return {
            'is_valid': self.is_valid,
            'error_count': len(self.errors),
            'warning_count': len(self.warnings),
            'errors': self.errors,
            'warnings': self.warnings
        }


class DataValidator:
    """
    데이터 검증을 위한 유틸리티 클래스
    """
    
    def __init__(self, logger: Optional[logging.Logger] = None):
        self.logger = logger or logging.getLogger(__name__)
    
    def validate_product_data(self, product_data: Dict[str, Any]) -> ValidationResult:
        """
        제품 데이터 검증
        
        Args:
            product_data: 제품 데이터 딕셔너리
        
        Returns:
            검증 결과
        """
        result = ValidationResult()
        
        # 필수 필드 검증
        required_fields = ['abbreviation', 'specification', 'manufacturing_number', 'batch_quantity']
        for field in required_fields:
            if not product_data.get(field):
                result.add_error(f"필수 필드 '{field}'가 비어있습니다")
        
        # 제품 코드 검증
        if product_data.get('abbreviation'):
            if not self.validate_product_code(product_data['abbreviation']):
                result.add_error("제품 코드 형식이 올바르지 않습니다")
        
        # 배치 수량 검증
        batch_quantity = product_data.get('batch_quantity')
        if batch_quantity is not None:
            if not isinstance(batch_quantity, (int, float)) or batch_quantity <= 0:
                result.add_error("배치 수량은 0보다 큰 숫자여야 합니다")
        
        # 우선순위 검증
        priority = product_data.get('priority')
        if priority is not None:
            if not isinstance(priority, int) or not (1 <= priority <= 10):
                result.add_error("우선순위는 1-10 사이의 정수여야 합니다")
        
        # 리드타임 검증
        lead_time = product_data.get('lead_time')
        if lead_time is not None:
            if not isinstance(lead_time, (int, float)) or lead_time < 0:
                result.add_error("리드타임은 0 이상의 숫자여야 합니다")
        
        return result
    
    def validate_schedule_data(self, schedule_data: Dict[str, Any]) -> ValidationResult:
        """
        스케줄 데이터 검증
        
        Args:
            schedule_data: 스케줄 데이터 딕셔너리
        
        Returns:
            검증 결과
        """
        result = ValidationResult()
        
        # 필수 필드 검증
        required_fields = ['schedule_id', 'schedule_name', 'start_date', 'end_date']
        for field in required_fields:
            if not schedule_data.get(field):
                result.add_error(f"필수 필드 '{field}'가 비어있습니다")
        
        # 날짜 검증
        start_date = schedule_data.get('start_date')
        end_date = schedule_data.get('end_date')
        
        if start_date and end_date:
            if isinstance(start_date, str):
                start_date = self.parse_date(start_date)
            if isinstance(end_date, str):
                end_date = self.parse_date(end_date)
            
            if start_date and end_date:
                if start_date >= end_date:
                    result.add_error("시작일은 종료일보다 빠를야 합니다")
        
        return result
    
    def validate_excel_structure(self, file_path: Union[str, Path], 
                               required_sheets: List[str]) -> ValidationResult:
        """
        엑셀 파일 구조 검증
        
        Args:
            file_path: 엑셀 파일 경로
            required_sheets: 필수 시트 목록
        
        Returns:
            검증 결과
        """
        result = ValidationResult()
        
        try:
            file_path = Path(file_path)
            
            # 파일 존재 여부 검증
            if not file_path.exists():
                result.add_error(f"파일을 찾을 수 없습니다: {file_path}")
                return result
            
            # 파일 확장자 검증
            if file_path.suffix.lower() not in ['.xlsx', '.xls']:
                result.add_error("엑셀 파일(.xlsx, .xls)만 지원됩니다")
                return result
            
            # 엑셀 파일 열기 및 시트 검증
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            existing_sheets = set(wb.sheetnames)
            required_sheets_set = set(required_sheets)
            
            missing_sheets = required_sheets_set - existing_sheets
            if missing_sheets:
                result.add_error(f"필요한 시트가 누락되었습니다: {', '.join(missing_sheets)}")
            
            # 각 시트의 데이터 여부 검증
            for sheet_name in required_sheets:
                if sheet_name in existing_sheets:
                    ws = wb[sheet_name]
                    if ws.max_row <= 1:  # 헤더만 있는 경우
                        result.add_warning(f"시트 '{sheet_name}'에 데이터가 없습니다")
            
            wb.close()
            
        except Exception as e:
            result.add_error(f"엑셀 파일 검증 중 오류 발생: {str(e)}")
        
        return result
    
    def validate_dataframe(self, df: pd.DataFrame, 
                          required_columns: List[str],
                          column_types: Optional[Dict[str, type]] = None) -> ValidationResult:
        """
        DataFrame 검증
        
        Args:
            df: 검증할 DataFrame
            required_columns: 필수 컬럼 목록
            column_types: 컬럼별 예상 타입
        
        Returns:
            검증 결과
        """
        result = ValidationResult()
        
        # 빈 DataFrame 검증
        if df.empty:
            result.add_error("DataFrame이 비어있습니다")
            return result
        
        # 필수 컬럼 검증
        existing_columns = set(df.columns)
        required_columns_set = set(required_columns)
        missing_columns = required_columns_set - existing_columns
        
        if missing_columns:
            result.add_error(f"필요한 컬럼이 누락되었습니다: {', '.join(missing_columns)}")
        
        # 컬럼 타입 검증
        if column_types:
            for column, expected_type in column_types.items():
                if column in df.columns:
                    if not df[column].dtype.type == expected_type:
                        result.add_warning(f"컬럼 '{column}'의 타입이 예상과 다릅니다. 예상: {expected_type.__name__}, 실제: {df[column].dtype}")
        
        # 누락값 검증
        for column in required_columns:
            if column in df.columns:
                null_count = df[column].isnull().sum()
                if null_count > 0:
                    result.add_warning(f"컬럼 '{column}'에 {null_count}개의 누락값이 있습니다")
        
        return result
    
    def validate_product_code(self, code: str) -> bool:
        """
        제품 코드 형식 검증
        
        Args:
            code: 제품 코드
        
        Returns:
            유효성 여부
        """
        if not isinstance(code, str) or len(code) == 0:
            return False
        
        # 기본 패턴: 영문자와 숫자, 하이픈만 허용
        pattern = r'^[A-Za-z0-9\-_]{3,20}$'
        return bool(re.match(pattern, code))
    
    def validate_email(self, email: str) -> bool:
        """
        이메일 주소 검증
        
        Args:
            email: 이메일 주소
        
        Returns:
            유효성 여부
        """
        if not isinstance(email, str):
            return False
        
        pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return bool(re.match(pattern, email))
    
    def validate_phone_number(self, phone: str) -> bool:
        """
        전화번호 검증 (한국 형식)
        
        Args:
            phone: 전화번호
        
        Returns:
            유효성 여부
        """
        if not isinstance(phone, str):
            return False
        
        # 한국 전화번호 패턴 (010-xxxx-xxxx, 02-xxxx-xxxx 등)
        patterns = [
            r'^010-\d{4}-\d{4}$',
            r'^0\d{1,2}-\d{3,4}-\d{4}$',
            r'^\d{2,3}-\d{3,4}-\d{4}$'
        ]
        
        return any(re.match(pattern, phone) for pattern in patterns)
    
    def parse_date(self, date_str: str) -> Optional[datetime]:
        """
        날짜 문자열 파싱
        
        Args:
            date_str: 날짜 문자열
        
        Returns:
            파싱된 datetime 객체 또는 None
        """
        if not isinstance(date_str, str):
            return None
        
        date_formats = [
            '%Y-%m-%d',
            '%Y/%m/%d',
            '%Y-%m-%d %H:%M:%S',
            '%Y/%m/%d %H:%M:%S',
            '%m/%d/%Y',
            '%d/%m/%Y'
        ]
        
        for fmt in date_formats:
            try:
                return datetime.strptime(date_str.strip(), fmt)
            except ValueError:
                continue
        
        return None
    
    def validate_range(self, value: Union[int, float], min_val: Union[int, float], 
                      max_val: Union[int, float]) -> bool:
        """
        숫자 범위 검증
        
        Args:
            value: 검증할 값
            min_val: 최소값
            max_val: 최대값
        
        Returns:
            범위 내 여부
        """
        if not isinstance(value, (int, float)):
            return False
        
        return min_val <= value <= max_val
    
    def validate_file_size(self, file_path: Union[str, Path], 
                          max_size_mb: float = 100) -> ValidationResult:
        """
        파일 크기 검증
        
        Args:
            file_path: 파일 경로
            max_size_mb: 최대 크기 (MB)
        
        Returns:
            검증 결과
        """
        result = ValidationResult()
        
        try:
            file_path = Path(file_path)
            
            if not file_path.exists():
                result.add_error(f"파일을 찾을 수 없습니다: {file_path}")
                return result
            
            file_size_mb = file_path.stat().st_size / (1024 * 1024)
            
            if file_size_mb > max_size_mb:
                result.add_error(f"파일 크기가 너무 큽니다: {file_size_mb:.2f}MB (max: {max_size_mb}MB)")
            
        except Exception as e:
            result.add_error(f"파일 크기 검증 중 오류: {str(e)}")
        
        return result
    
    def create_custom_validator(self, validation_func: Callable[[Any], bool], 
                              error_message: str) -> Callable[[Any], ValidationResult]:
        """
        사용자 정의 검증자 생성
        
        Args:
            validation_func: 검증 함수
            error_message: 오류 메시지
        
        Returns:
            검증자 함수
        """
        def validator(value: Any) -> ValidationResult:
            result = ValidationResult()
            try:
                if not validation_func(value):
                    result.add_error(error_message)
            except Exception as e:
                result.add_error(f"검증 중 오류 발생: {str(e)}")
            return result
        
        return validator


# 편의 함수들
def validate_product(product_data: Dict[str, Any]) -> ValidationResult:
    """
    제품 데이터 간단 검증
    
    Args:
        product_data: 제품 데이터
    
    Returns:
        검증 결과
    """
    validator = DataValidator()
    return validator.validate_product_data(product_data)


def validate_excel(file_path: Union[str, Path], required_sheets: List[str]) -> ValidationResult:
    """
    엑셀 파일 간단 검증
    
    Args:
        file_path: 파일 경로
        required_sheets: 필수 시트
    
    Returns:
        검증 결과
    """
    validator = DataValidator()
    return validator.validate_excel_structure(file_path, required_sheets)


def is_valid_product_code(code: str) -> bool:
    """
    제품 코드 유효성 간단 검증
    
    Args:
        code: 제품 코드
    
    Returns:
        유효성 여부
    """
    validator = DataValidator()
    return validator.validate_product_code(code)


def is_valid_email(email: str) -> bool:
    """
    이메일 주소 유효성 간단 검증
    
    Args:
        email: 이메일 주소
    
    Returns:
        유효성 여부
    """
    validator = DataValidator()
    return validator.validate_email(email)

