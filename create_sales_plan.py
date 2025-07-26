"""
판매계획 더미데이터 생성 스크립트
"""
import pandas as pd
import numpy as np
from datetime import datetime
import os

# 판매계획 데이터 생성
def create_sales_plan():
    # 제품 목록 (마스터 데이터와 일치)
    products = ['제품A', '제품B', '제품C']
    
    # 월별 판매계획 데이터 생성
    # 제품A: 우선순위 높음, 수요 많음
    # 제품B: 중간 우선순위, 중간 수요
    # 제품C: 낮은 우선순위, 적은 수요
    
    sales_data = {
        '제품명': products,
        '1월': [120, 80, 50],
        '2월': [130, 85, 55],
        '3월': [140, 90, 60],
        '4월': [135, 88, 58],
        '5월': [145, 95, 65],
        '6월': [150, 100, 70],
        '7월': [155, 105, 75],
        '8월': [160, 110, 80],
        '9월': [150, 105, 75],
        '10월': [145, 100, 70],
        '11월': [140, 95, 65],
        '12월': [135, 90, 60]
    }
    
    # DataFrame 생성
    df = pd.DataFrame(sales_data)
    
    # 연간 총계 컬럼 추가
    month_columns = [col for col in df.columns if '월' in col]
    df['연간총계'] = df[month_columns].sum(axis=1)
    
    # Excel 파일로 저장
    output_path = os.path.join('data', 'sales_plan_2025.xlsx')
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='판매계획', index=False)
        
        # 워크시트 가져오기
        worksheet = writer.sheets['판매계획']
        
        # 헤더 스타일 적용
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color='1428A0', end_color='1428A0', fill_type='solid')
        header_font_white = Font(bold=True, color='FFFFFF')
        
        # 헤더 행 스타일
        for cell in worksheet[1]:
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # 컬럼 너비 조정
        worksheet.column_dimensions['A'].width = 15
        for col in worksheet.columns:
            if col[0].column > 1:
                worksheet.column_dimensions[col[0].column_letter].width = 10
        
        # 데이터 정렬
        for row in worksheet.iter_rows(min_row=2):
            row[0].alignment = Alignment(horizontal='left')
            for cell in row[1:]:
                cell.alignment = Alignment(horizontal='center')
    
    print(f"판매계획 파일이 생성되었습니다: {output_path}")
    
    # 추가 정보 생성 (메타데이터)
    metadata = {
        '작성일': datetime.now().strftime('%Y-%m-%d'),
        '작성자': 'APS System',
        '단위': '배치(Batch)',
        '비고': '2025년 연간 판매계획'
    }
    
    # 요약 정보
    summary_data = []
    for idx, product in enumerate(products):
        summary_data.append({
            '제품명': product,
            '월평균': round(df.iloc[idx][month_columns].mean(), 1),
            '최대': df.iloc[idx][month_columns].max(),
            '최소': df.iloc[idx][month_columns].min(),
            '연간총계': df.iloc[idx]['연간총계']
        })
    
    summary_df = pd.DataFrame(summary_data)
    
    # 상세 판매계획 파일 생성 (여러 시트)
    output_detail_path = os.path.join('data', 'sales_plan_2025_detail.xlsx')
    
    with pd.ExcelWriter(output_detail_path, engine='openpyxl') as writer:
        # 메인 판매계획
        df.to_excel(writer, sheet_name='판매계획', index=False)
        
        # 요약 정보
        summary_df.to_excel(writer, sheet_name='요약', index=False)
        
        # 분기별 집계
        quarterly_data = {
            '제품명': products,
            '1분기': [df.iloc[i][['1월', '2월', '3월']].sum() for i in range(len(products))],
            '2분기': [df.iloc[i][['4월', '5월', '6월']].sum() for i in range(len(products))],
            '3분기': [df.iloc[i][['7월', '8월', '9월']].sum() for i in range(len(products))],
            '4분기': [df.iloc[i][['10월', '11월', '12월']].sum() for i in range(len(products))]
        }
        quarterly_df = pd.DataFrame(quarterly_data)
        quarterly_df.to_excel(writer, sheet_name='분기별', index=False)
        
    print(f"상세 판매계획 파일이 생성되었습니다: {output_detail_path}")
    
    return df

if __name__ == "__main__":
    create_sales_plan()